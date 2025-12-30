import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
import os
import sys
import argparse

def unmerge_cells_and_fill(file_path):
    """Load Excel, unmerge cells, and fill downward"""
    
    # Load workbook with openpyxl
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Create a copy of merged cells ranges before unmerging
    merged_cells = list(ws.merged_cells.ranges)
    
    # First, store all values from merged cells
    merged_values = {}
    for merged_range in merged_cells:
        # Get the value from the top-left cell
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        top_left_value = top_left_cell.value
        
        # Store the value for all cells in this merged range
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell_ref = (row, col)
                merged_values[cell_ref] = top_left_value
    
    # Now unmerge all cells
    for merged_range in merged_cells:
        ws.unmerge_cells(str(merged_range))
    
    # Apply the stored values to all previously merged cells
    for (row, col), value in merged_values.items():
        # Get the cell (now unmerged) and set its value
        cell = ws.cell(row=row, column=col)
        cell.value = value
    
    # Save to a temporary file
    temp_file = "temp_unmerged.xlsx"
    wb.save(temp_file)
    
    # Read with pandas
    df = pd.read_excel(temp_file, engine='openpyxl')
    
    # Optional: Clean up temp file
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    return df, wb

def simple_fill_blanks(file_path):
    """Simple method - just forward fill blank cells"""
    
    # Read with pandas
    df = pd.read_excel(file_path, engine='openpyxl')
    
    # Forward fill to fill blank cells with previous values
    df_filled = df.ffill()
    
    return df_filled

def save_to_excel_and_json(df, base_filename, output_dir="output"):
    """Save DataFrame to both Excel and JSON"""
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Create base filenames
    excel_output = os.path.join(output_dir, f"{base_filename}_unmerged.xlsx")
    json_output = os.path.join(output_dir, f"{base_filename}.json")
    
    # Save to Excel for verification
    df.to_excel(excel_output, index=False)
    print(f"✓ Unmerged data saved to: {excel_output}")
    
    # Convert to JSON
    json_data = df.to_dict(orient='records')
    
    # Save to JSON with proper formatting
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    
    print(f"✓ JSON data saved to: {json_output}")
    
    return excel_output, json_output

def process_file(input_file, method='simple', output_dir="output"):
    """Process a single Excel file"""
    
    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found!")
        return None
    
    print(f"\n{'='*60}")
    print(f"Processing: {input_file}")
    print(f"{'='*60}")
    
    # Get base filename without extension
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    
    try:
        if method == 'simple':
            print("Using simple forward-fill method...")
            df = simple_fill_blanks(input_file)
            excel_file, json_file = save_to_excel_and_json(df, base_name, output_dir)
            
        elif method == 'unmerge':
            print("Using unmerge method...")
            df, _ = unmerge_cells_and_fill(input_file)
            excel_file, json_file = save_to_excel_and_json(df, base_name, output_dir)
        
        elif method == 'both':
            print("Trying both methods...")
            
            # Simple method
            df_simple = simple_fill_blanks(input_file)
            excel_simple, json_simple = save_to_excel_and_json(
                df_simple, f"{base_name}_simple", output_dir
            )
            
            # Unmerge method
            try:
                df_unmerged, _ = unmerge_cells_and_fill(input_file)
                excel_unmerged, json_unmerged = save_to_excel_and_json(
                    df_unmerged, f"{base_name}_unmerged", output_dir
                )
            except Exception as e:
                print(f"⚠ Unmerge method failed: {e}")
        
        else:
            print(f"❌ Unknown method: {method}")
            return None
        
        # Show preview
        print("\n📊 Preview of data (first 5 rows):")
        print(df.head())
        print(f"\n📋 Column names: {df.columns.tolist()}")
        print(f"📈 Total rows: {len(df)}")
        print(f"📊 Total columns: {len(df.columns)}")
        
        return df
        
    except Exception as e:
        print(f"❌ Error processing {input_file}: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    parser = argparse.ArgumentParser(
        description='Convert Excel files with merged cells to JSON',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s data.xlsx                    # Process single file with simple method
  %(prog)s file1.xlsx file2.xlsx        # Process multiple files
  %(prog)s *.xlsx                       # Process all Excel files
  %(prog)s data.xlsx -m unmerge         # Use unmerge method
  %(prog)s data.xlsx -o custom_output   # Specify output directory
  %(prog)s data.xlsx -p                 # Show preview only
        """
    )
    
    parser.add_argument(
        'files', 
        nargs='+',
        help='Excel file(s) to process'
    )
    
    parser.add_argument(
        '-m', '--method',
        choices=['simple', 'unmerge', 'both'],
        default='simple',
        help='Processing method (default: simple)'
    )
    
    parser.add_argument(
        '-o', '--output',
        default='output',
        help='Output directory (default: output/)'
    )
    
    parser.add_argument(
        '-p', '--preview',
        action='store_true',
        help='Show preview only, don\'t save files'
    )
    
    parser.add_argument(
        '--version',
        action='version',
        version='Excel to JSON Converter 1.0'
    )
    
    args = parser.parse_args()
    
    print("🔧 Excel to JSON Converter")
    print(f"📁 Input files: {len(args.files)}")
    print(f"⚙ Method: {args.method}")
    print(f"📂 Output directory: {args.output}")
    print()
    
    # Process each file
    successful = 0
    for file_path in args.files:
        if args.preview:
            # Preview mode
            if os.path.exists(file_path):
                try:
                    df = simple_fill_blanks(file_path)
                    print(f"\n📄 File: {file_path}")
                    print(f"📊 Shape: {df.shape[0]} rows × {df.shape[1]} columns")
                    print("\nFirst 3 rows:")
                    print(df.head(3))
                    print("\nColumn names:")
                    print(df.columns.tolist())
                    print("-" * 40)
                    successful += 1
                except Exception as e:
                    print(f"❌ Error previewing {file_path}: {e}")
            else:
                print(f"❌ File not found: {file_path}")
        else:
            # Process and save mode
            result = process_file(file_path, args.method, args.output)
            if result is not None:
                successful += 1
    
    print(f"\n{'='*60}")
    print(f"✅ Processing complete!")
    print(f"   Successfully processed: {successful}/{len(args.files)} files")
    print(f"   Output directory: {args.output}")
    
    if not args.preview and successful > 0:
        print("\n📁 Files created:")
        if os.path.exists(args.output):
            for file in sorted(os.listdir(args.output)):
                if file.endswith('.json') or file.endswith('.xlsx'):
                    size = os.path.getsize(os.path.join(args.output, file))
                    print(f"   • {file} ({size:,} bytes)")

if __name__ == "__main__":
    main()